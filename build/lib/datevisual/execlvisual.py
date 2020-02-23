import xlrd
import openpyxl
import json
import os
from pyecharts import Graph

'''
execlpath 文件路径
Preservationpath 保存路径 保存文件名称为导入的execl文件名+收付标志，类型为html
sourceaccount 原账户 输入属性名称 例如 '交易账卡号'
targetaccount 目的账户 输入属性名称 '对手账号'
value 权重 输入属性名称
label='Flase'  是否有收付标志，默认没有  如果有输入属性名称
'''

#支持xls和xlsx两种格式
def execlfileVisual(execlpath, Preservationpath,sourceaccount,targetaccount,value,label='Flase'):
    fileFormat = os.path.splitext(execlpath)[1]
    if fileFormat == '.xls':       
        xlsVisual(execlpath, Preservationpath,sourceaccount,targetaccount,value,label)
    elif fileFormat == '.xlsx':       
        xlsxVisual(execlpath, Preservationpath,sourceaccount,targetaccount,value,label)
    else:
        print("不支持此格式文件！")




def xlsVisual(execlpath, Preservationpath, sourceaccount, targetaccount, value, label):
    if label == 'Flase':        
        xlsNolabelVisucal(execlpath, Preservationpath, sourceaccount, targetaccount, value)
    else:       
        xlsHavelabelVisucal(execlpath, Preservationpath, sourceaccount, targetaccount, value , label)
    


def xlsxVisual(execlpath, Preservationpath,sourceaccount,targetaccount,value,label):
    if label == 'Flase':       
        xlsxNolabelVisucal(execlpath, Preservationpath, sourceaccount, targetaccount, value)
    else:       
        xlsxHavelabelVisucal(execlpath, Preservationpath, sourceaccount, targetaccount, value , label)


#xls 无收付标志可视化
def xlsNolabelVisucal(execlpath, Preservationpath, sourceaccount, targetaccount, value):
    accountSet = set()
    linksList=[]
    nodeList=[]
    workbook = xlrd.open_workbook(execlpath)
    sheets = workbook.sheet_names()
    for sheetnum in sheets:
        worksheet = workbook.sheet_by_name(sheetnum)
        firstrowvalue = worksheet.row_values(0)
        print(firstrowvalue)

        #查找原账户位置
        if sourceaccount in firstrowvalue:
            sourceIndex = firstrowvalue.index(sourceaccount)
        else:
            print("文件中没有%s属性", sourceaccount)
            break
        #查找目的账户位置
        if targetaccount in firstrowvalue:
            targetIndex = firstrowvalue.index(targetaccount)
        else:
            print("文件中没有%s属性", targetaccount)
            break         
        #查找权值位置
        if value in firstrowvalue:
            valueIndex = firstrowvalue.index(value)
        else:
            print("文件中没有%s属性", value)
            break

        for i in range(1,worksheet.nrows):
            onerowdict={}
            sourcenode = worksheet.cell_value(i, sourceIndex)
            targetnode = worksheet.cell_value(i, targetIndex)
            valuenode = float(worksheet.cell_value(i, valueIndex))

            if sourcenode == '' or targetnode == '':
                continue

            accountSet.add(sourcenode)
            accountSet.add(targetnode)
            if len(linksList) > 0:
                flog=0
                for x in linksList:
                    if sourcenode == x['source'] and targetnode == x['target']:
                        x['value'] += valuenode
                        flog = 1
                if flog == 0:
                    onerowdict['source'] = sourcenode
                    onerowdict['target'] = targetnode
                    onerowdict['value'] = valuenode
                    linksList.append(onerowdict)
            else:
                onerowdict['source'] = sourcenode
                onerowdict['target'] = targetnode
                onerowdict['value'] = valuenode
                linksList.append(onerowdict)   

    for node in accountSet:
        nodeDict = {}
        nodeDict['name'] = node
        nodeDict['symbolSize']=15
        nodeList.append(nodeDict)
    nodes = nodeList
    links = linksList
    
    print('节点数量：',len(nodes))
    print('边数量：',len(links))

    echartshow(nodes, links, execlpath, Preservationpath, flogg=False)

    
#xls 有收付标志可视化
def xlsHavelabelVisucal(execlpath, Preservationpath, sourceaccount, targetaccount, value, label):
    accountSet = set()
    linksList = []
    nodeList=[]
    workbook = xlrd.open_workbook(execlpath)
    sheets = workbook.sheet_names()
    for sheetnum in sheets:
        worksheet = workbook.sheet_by_name(sheetnum)
        firstrowvalue = worksheet.row_values(0)
        print(firstrowvalue)
        #查找原账户位置
        if sourceaccount in firstrowvalue:
            sourceIndex = firstrowvalue.index(sourceaccount)
        else:
            print("文件中没有%s属性", sourceaccount)
            break
        #查找目的账户位置
        if targetaccount in firstrowvalue:
            targetIndex = firstrowvalue.index(targetaccount)
        else:
            print("文件中没有%s属性", targetaccount)
            break         
        #查找权值位置
        if value in firstrowvalue:
            valueIndex = firstrowvalue.index(value)
        else:
            print("文件中没有%s属性", value)
            break
        #查找标签位置
        if label in firstrowvalue:
                labelIndex = firstrowvalue.index(label)
        else:
            print("文件中没有%s属性", label)
            break

        for i in range(1,worksheet.nrows):
            onerowdict = {}
            sourcenode = worksheet.cell_value(i, sourceIndex)
            targetnode = worksheet.cell_value(i, targetIndex)
            valuenode = float(worksheet.cell_value(i, valueIndex))

            if sourcenode == '' or targetnode == '':
                continue

            accountSet.add(sourcenode)
            accountSet.add(targetnode)

            if worksheet.cell_value(i, labelIndex) == '出':
                
                if len(linksList) > 0:
                    flog=0
                    for x in linksList:
                        if sourcenode == x['source'] and targetnode == x['target']:
                            x['value'] += valuenode
                            flog = 1
                    if flog == 0:
                        onerowdict['source'] = sourcenode
                        onerowdict['target'] = targetnode
                        onerowdict['value'] = valuenode
                        linksList.append(onerowdict)
                else:
                    onerowdict['source'] = sourcenode
                    onerowdict['target'] = targetnode
                    onerowdict['value'] = valuenode
                    linksList.append(onerowdict)
                        
            if worksheet.cell_value(i, labelIndex) == '进':
                
                if len(linksList) > 0:
                    flog=0
                    for x in linksList:
                        if targetnode == x['source'] and sourcenode == x['target']:
                            x['value'] += valuenode
                            flog = 1
                    if flog == 0:
                        onerowdict['source'] = targetnode
                        onerowdict['target'] = sourcenode
                        onerowdict['value'] = valuenode
                        linksList.append(onerowdict)
                else:
                    onerowdict['source'] = targetnode
                    onerowdict['target'] = sourcenode
                    onerowdict['value'] = valuenode
                    linksList.append(onerowdict)
          
    for node in accountSet:
        nodeDict = {}
        nodeDict['name'] = node
        nodeDict['symbolSize']=15
        nodeList.append(nodeDict)
    nodes = nodeList
    links = linksList
    
    print('节点数量：',len(nodes))
    print('边数量：', len(links))
    
    echartshow(nodes, links, execlpath, Preservationpath, flogg=True)


   
#xlsx 无收付标志可视化
def xlsxNolabelVisucal(execlpath, Preservationpath, sourceaccount, targetaccount, value):
    accountSet = set()
    linksList = []
    nodeList = []
    firstrowvalue=[]
    workbook = openpyxl.load_workbook(execlpath)
    sheets = workbook.sheetnames
    for sheetnum in sheets:
        worksheet = workbook[sheetnum]
        for cell in list(worksheet.rows)[0]:
            firstrowvalue.append(cell.value)
        print(firstrowvalue)
        #查找原账户位置
        if sourceaccount in firstrowvalue:
            sourceIndex = firstrowvalue.index(sourceaccount)+1
        else:
            print("文件中没有%s属性", sourceaccount)
            break
        #查找目的账户位置
        if targetaccount in firstrowvalue:
            targetIndex = firstrowvalue.index(targetaccount)+1
        else:
            print("文件中没有%s属性", targetaccount)
            break         
        #查找权值位置
        if value in firstrowvalue:
            valueIndex = firstrowvalue.index(value)+1
        else:
            print("文件中没有%s属性", value)
            break
        
        for i in range(2, worksheet.max_row):
            onerowdict = {}
            sourcenode= worksheet.cell(row=i, column=sourceIndex).value
            targetnode = worksheet.cell(row=i, column=targetIndex).value
            valuenode= float(worksheet.cell(row=i, column=valueIndex).value)

            if sourcenode == '' or targetnode == '' or sourcenode == None or targetnode == None:
                continue

            accountSet.add(sourcenode)
            accountSet.add(targetnode)
            if len(linksList) > 0:
                flog=0
                for x in linksList:
                    if sourcenode == x['source'] and targetnode == x['target']:
                        x['value'] += valuenode
                        flog = 1
                if flog == 0:
                    onerowdict['source'] = sourcenode
                    onerowdict['target'] = targetnode
                    onerowdict['value'] = valuenode
                    linksList.append(onerowdict)
            else:
                onerowdict['source'] = sourcenode
                onerowdict['target'] = targetnode
                onerowdict['value'] = valuenode
                linksList.append(onerowdict)

    
    for node in accountSet:
        nodeDict = {}
        nodeDict['name'] = node
        nodeDict['symbolSize']=15
        nodeList.append(nodeDict)
    nodes = nodeList
    links = linksList
    
    print('节点数量：',len(nodes))
    print('边数量：', len(links))
    
    echartshow(nodes, links, execlpath, Preservationpath, flogg=False)

    
#xlsx 有收付标志可视化
def xlsxHavelabelVisucal(execlpath, Preservationpath, sourceaccount, targetaccount, value , label):
    accountSet = set()
    linksList = []
    nodeList = []
    firstrowvalue=[]
    workbook = openpyxl.load_workbook(execlpath)
    sheets = workbook.sheetnames
    for sheetnum in sheets:
        worksheet = workbook[sheetnum]

        for cell in list(worksheet.rows)[0]:
            firstrowvalue.append(cell.value)
        print(firstrowvalue)
        #查找原账户位置
        if sourceaccount in firstrowvalue:
            sourceIndex = firstrowvalue.index(sourceaccount)+1
        else:
            print("文件中没有%s属性", sourceaccount)
            break
        #查找目的账户位置
        if targetaccount in firstrowvalue:
            targetIndex = firstrowvalue.index(targetaccount)+1
        else:
            print("文件中没有%s属性", targetaccount)
            break         
        #查找权值位置
        if value in firstrowvalue:
            valueIndex = firstrowvalue.index(value)+1
        else:
            print("文件中没有%s属性", value)
            break
        #查找标签位置
        if label in firstrowvalue:
                labelIndex = firstrowvalue.index(label)+1
        else:
            print("文件中没有%s属性", label)
            break
        
        for i in range(2, worksheet.max_row):
            onerowdict = {}
            sourcenode= worksheet.cell(row=i, column=sourceIndex).value
            targetnode = worksheet.cell(row=i, column=targetIndex).value
            valuenode= float(worksheet.cell(row=i, column=valueIndex).value)

            if sourcenode == '' or targetnode == ''or sourcenode == None or targetnode == None:
                continue

            accountSet.add(sourcenode)
            accountSet.add(targetnode)

            if worksheet.cell(row=i, column=labelIndex).value == '出':
                
                if len(linksList) > 0:
                    flog=0
                    for x in linksList:
                        if sourcenode == x['source'] and targetnode == x['target']:
                            x['value'] += valuenode
                            flog=1
                    if flog==0:
                        onerowdict['source'] = sourcenode
                        onerowdict['target'] = targetnode
                        onerowdict['value'] = valuenode
                        linksList.append(onerowdict)
                else:
                    onerowdict['source'] = sourcenode
                    onerowdict['target'] = targetnode
                    onerowdict['value'] = valuenode
                    linksList.append(onerowdict)
            if worksheet.cell(row=i, column=labelIndex).value == '进':
                
                if len(linksList) > 0:
                    flog=0
                    for x in linksList:
                        if targetnode == x['source'] and sourcenode == x['target']:
                            x['value'] += valuenode
                            flog=1
                    if flog==0:
                        onerowdict['source'] = targetnode
                        onerowdict['target'] = sourcenode
                        onerowdict['value'] = valuenode
                        linksList.append(onerowdict)
                else:
                    onerowdict['source'] = targetnode
                    onerowdict['target'] = sourcenode
                    onerowdict['value'] = valuenode
                    linksList.append(onerowdict)

    for node in accountSet:
        nodeDict = {}
        nodeDict['name'] = node
        nodeDict['symbolSize']=15
        nodeList.append(nodeDict)
    nodes = nodeList
    links = linksList
    
    print('节点数量：',len(nodes))
    print('边数量：', len(links))
    
    echartshow(nodes, links, execlpath, Preservationpath, flogg=True)



def echartshow(nodes, links, execlpath, Preservationpath, flogg=False):
    if flogg:
        partstr = '有收付标志'
    else:
        partstr = '无收付标志'
      
    graph = Graph(os.path.basename(execlpath).split('.')[0]+partstr+"图例",width=1920, height=1080)
    graph.add(
        "", 
        nodes, 
        links, 
        is_label_show=False,
        repulsion=50,
        is_focusnode=True,
        is_roam=True,
        graph_layout='force',
        line_color="rgba（50,50,50,0.7）",
        graph_edge_symbol= ['circle', 'arrow']
        )
    graph.render((str(Preservationpath)+os.path.basename(execlpath).split('.')[0]+partstr+'.html'))



# if __name__ == "__main__":
#     execlpath = r'账户交易明细表.xls'
#     execlpath1=r'账户交易明细表副本.xlsx'
#     Preservationpath = r'result/'
#     sourceaccount = '交易账卡号' 
#     targetaccount = '对手账号'
#     value = '交易金额'
#     label = '收付标志'
#     #xls
#     execlfileVisual(execlpath, Preservationpath, sourceaccount, targetaccount, value, label)
#     execlfileVisual(execlpath, Preservationpath, sourceaccount, targetaccount, value)
#     # #xlsx
#     execlfileVisual(execlpath1, Preservationpath, sourceaccount, targetaccount, value, label)
#     execlfileVisual(execlpath1, Preservationpath, sourceaccount, targetaccount, value)